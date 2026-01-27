"""
CWE Classification Metrics Calculator (Scenario 2: Validated Exact Match)

This script calculates classification metrics based on an exact match between the
predicted CWEs and the ground truth. It is designed to be a robust tool for
evaluating model performance with a strict correctness criterion.

Key Features:
- Standardizes all CWE tags (e.g., 'CWE-022' -> 'CWE-22') to prevent mismatches
  due to formatting inconsistencies.
- Validates predicted CWEs against the official MITRE CWE list, discarding any
  invalid identifiers.
- Ignores predicted CWEs that do not exist in the ground truth vocabulary of the dataset.
- Calculates a focused set of per-class, macro, and weighted average metrics
  (TP, FP, FN, TN, F1-Score, FPR, FNR).
- Operates in two modes: scanning a full directory of models or analyzing a single file.

MODES OF OPERATION:
1. Directory Mode (default):
   - Scans a base directory for model subfolders.
   - For each model, processes Sven/PrimeVul/DiverseVul/parser_output_prompt_i folders.
   - Usage: python scenario2_final_mod.py /path/to/models_directory

2. Single File Mode:
   - Processes only one specified Excel file.
   - Expects the file under Model/Dataset/parser_output_prompt_X.
   - Saves the report in Dataset/metrics-scenario-2.
   - Usage: python scenario2_final_mod.py -f /path/to/Models/Model_A/Sven/parser_output_prompt_1/file.xlsx
"""

# --- Core Libraries ---
import pandas as pd
import argparse
import re
import numpy as np
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Set
import xml.etree.ElementTree as ET
import io
import zipfile
import sys

# --- Optional Libraries ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "--> WARNING: 'openpyxl' not found. Excel formatting will be disabled. Install with: pip install openpyxl"
    )
    openpyxl = None
    get_column_letter = None

try:
    import requests
except ImportError:
    print(
        "--> FATAL ERROR: 'requests' not found. It is required for XML download. Install with: pip install requests"
    )
    requests = None

# --- CONFIGURATION CONSTANTS ---
COL_FILENAME = "File Name"
COL_PREDICTED = "Found CWE"
COL_GROUND_TRUTH = "Actual CWE"
CWE_LATEST_URL = "https://cwe.mitre.org/data/xml/cwec_latest.xml.zip"

# Dataset attesi nella struttura Models/Model_X/<Dataset>/parser_output_prompt_X
DATASETS = ["Sven", "PrimeVul", "DiverseVul"]


def ensure_cwe_xml_exists(xml_path: Path) -> bool:
    """
    Checks if the CWE XML file exists. If not, it attempts to download and extract it.
    """
    if xml_path.is_file():
        print(f"  -> Found existing CWE XML file: {xml_path}")
        return True
    if not requests:
        return False
    print(f"--- CWE XML file not found at '{xml_path}'. Attempting to download... ---")
    try:
        response = requests.get(CWE_LATEST_URL, timeout=60)
        response.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(response.content)) as z:
            xml_name = next(
                (name for name in z.namelist() if name.endswith(".xml")), None
            )
            if not xml_name:
                print("  --> FATAL ERROR: No XML file found in the downloaded archive.")
                return False
            print(f"  -> Extracting '{xml_name}' to '{xml_path}'...")
            xml_path.parent.mkdir(parents=True, exist_ok=True)
            with z.open(xml_name) as source, open(xml_path, "wb") as target:
                target.write(source.read())
        print("  -> Successfully downloaded CWE XML.")
        return True
    except Exception as e:
        print(f"  --> FATAL ERROR during XML download/extraction: {e}")
        return False


def parse_cwe_xml_for_ids(xml_file_path: Path) -> Optional[Set[str]]:
    """
    Parses the CWE XML to extract a set of all valid CWE IDs for validation.
    """
    print(f"  -> Parsing XML to get all valid CWE IDs from: {xml_file_path.name}")
    try:
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        ns = {"cwe": "http://cwe.mitre.org/cwe-7"}  # XML namespace is crucial
        cwe_ids = set()
        for element_type in ["Weaknesses", "Categories", "Views"]:
            container = root.find(f"cwe:{element_type}", ns)
            if container is not None:
                for item in container:
                    if item.get("ID"):
                        cwe_ids.add(item.get("ID"))
        if not cwe_ids:
            print("  --> WARNING: Could not extract any CWE IDs from the XML.")
            return None
        print(f"  -> Found {len(cwe_ids)} unique CWE entries in the XML.")
        return cwe_ids
    except (ET.ParseError, FileNotFoundError) as e:
        print(
            f"  --> FATAL ERROR: Could not read or parse XML file '{xml_file_path}'.\n      Details: {e}"
        )
        return None


def process_cwe_tags(series: pd.Series) -> List[str]:
    """
    Extracts, normalizes, standardizes, and deduplicates CWE tags from a pandas Series.
    """
    if series.dropna().empty:
        return []
    full_text = " ".join(series.dropna().astype(str))
    tags = re.findall(r"CWE-\d+|NOT VULNERABLE", full_text, re.IGNORECASE)

    standardized_tags = set()
    for tag in tags:
        tag_upper = tag.upper()
        if "CWE-" in tag_upper:
            try:
                cwe_number_str = re.search(r"\d+", tag_upper).group(0)
                standard_id = str(int(cwe_number_str))  # remove leading zeros
                standardized_tags.add(f"CWE-{standard_id}")
            except (AttributeError, ValueError):
                standardized_tags.add(tag_upper)
        else:
            standardized_tags.add(tag_upper)  # NOT VULNERABLE

    return sorted(list(standardized_tags))


def load_and_filter_data(
    file_path: Path, valid_cwe_ids_from_xml: Set[str]
) -> Optional[pd.DataFrame]:
    """
    Loads, aggregates, standardizes, and filters data from a single Excel file.
    """
    print(f"  -> Reading and aggregating file: {file_path.name}")
    try:
        df = pd.read_excel(file_path, engine="openpyxl")
        required_cols = {COL_FILENAME, COL_PREDICTED, COL_GROUND_TRUTH}
        if not required_cols.issubset(df.columns):
            print(
                f"  --> ERROR: Missing columns in {file_path.name}! Must contain: {', '.join(required_cols)}"
            )
            return None
    except Exception as e:
        print(
            f"  --> ERROR: Could not read the Excel file {file_path.name}. Details: {e}"
        )
        return None

    # Aggrega per file e normalizza le CWE
    aggregated_df = (
        df.groupby(COL_FILENAME)
        .agg({COL_GROUND_TRUTH: process_cwe_tags, COL_PREDICTED: process_cwe_tags})
        .reset_index()
    )

    # Filtra le Found CWE contro la lista ufficiale
    print("  -> Filtering 'Found CWE' against official XML data...")
    discard_report = []
    updated_predicted_cwe = []

    for _, row in aggregated_df.iterrows():
        filename, actual_cwes, original_predicted_cwes = (
            row[COL_FILENAME],
            row[COL_GROUND_TRUTH],
            row[COL_PREDICTED],
        )
        is_ground_truth_not_vulnerable = (
            len(actual_cwes) == 1 and actual_cwes[0] == "NOT VULNERABLE"
        )

        if is_ground_truth_not_vulnerable:
            updated_predicted_cwe.append(original_predicted_cwes)
        else:
            kept_cwes, discarded_cwes = [], []
            for tag in original_predicted_cwes:
                if tag.upper() == "NOT VULNERABLE":
                    kept_cwes.append(tag)
                    continue

                cwe_id_match = re.search(r"\d+", tag)
                if cwe_id_match:
                    cwe_id_str = cwe_id_match.group(0)
                    if cwe_id_str in valid_cwe_ids_from_xml:
                        kept_cwes.append(tag)
                    else:
                        discarded_cwes.append(tag)
                else:
                    discarded_cwes.append(tag)
            updated_predicted_cwe.append(kept_cwes)
            if discarded_cwes:
                discard_report.append(
                    {
                        "filename": filename,
                        "discarded": sorted(list(set(discarded_cwes))),
                    }
                )

    aggregated_df[COL_PREDICTED] = updated_predicted_cwe

    if discard_report:
        print("\n" + "-" * 50)
        print(
            f"  Summary of invalid CWEs discarded from 'Found CWE' in: {file_path.name}"
        )
        print("-" * 50)
        for item in discard_report:
            print(
                f"  - From row '{item['filename']}': Discarded: {', '.join(item['discarded'])}"
            )
        print("-" * 50 + "\n")
    return aggregated_df


def generate_metric_reports(
    df: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Calculates classification metrics (per-class, macro, and weighted averages).
    """
    print("  -> Calculating metrics...")
    actuals = df.explode(COL_GROUND_TRUTH).rename(columns={COL_GROUND_TRUTH: "Class"})
    predicted = df.explode(COL_PREDICTED).rename(columns={COL_PREDICTED: "Class"})
    if actuals["Class"].dropna().empty:
        print("  --> WARNING: No valid 'Actual CWE' tags found to calculate metrics.")
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

    # Core counts
    support = actuals["Class"].value_counts().rename("Support")
    tp_df = pd.merge(actuals, predicted, on=[COL_FILENAME, "Class"])
    tp = tp_df["Class"].value_counts().rename("TP")
    per_class_df = pd.concat([tp, support], axis=1).fillna(0).astype(np.int64)
    total_predictions = predicted["Class"].value_counts()
    per_class_df["FP"] = total_predictions.sub(per_class_df["TP"], fill_value=0).astype(
        np.int64
    )
    per_class_df["FN"] = per_class_df["Support"] - per_class_df["TP"]
    per_class_df["TN"] = len(df) - (
        per_class_df["TP"] + per_class_df["FP"] + per_class_df["FN"]
    )

    # Derived metrics
    with np.errstate(divide="ignore", invalid="ignore"):
        precision = (
            per_class_df["TP"] / (per_class_df["TP"] + per_class_df["FP"])
        ).fillna(0)
        recall = (per_class_df["TP"] / per_class_df["Support"]).fillna(0)
        per_class_df["F1-Score"] = (
            2 * precision * recall / (precision + recall)
        ).fillna(0)
        per_class_df["FPR"] = (
            per_class_df["FP"] / (per_class_df["FP"] + per_class_df["TN"])
        ).fillna(0)
        per_class_df["FNR"] = (per_class_df["FN"] / per_class_df["Support"]).fillna(0)

    avg_cols = ["F1-Score", "FPR", "FNR"]
    final_cols = ["TP", "FP", "FN", "TN", "Support"] + avg_cols
    per_class_report = per_class_df.reindex(columns=final_cols).sort_index().fillna(0)

    macro_avg = per_class_report[avg_cols].mean()
    weighted_avg = (
        per_class_report[avg_cols].multiply(per_class_report["Support"], axis=0)
    ).sum() / per_class_report["Support"].sum()

    return (
        per_class_report,
        pd.DataFrame(macro_avg).T.add_prefix("Macro "),
        pd.DataFrame(weighted_avg).T.add_prefix("Weighted "),
    )


def save_reports_to_excel(output_path: Path, reports: Dict[str, pd.DataFrame]):
    """
    Saves multiple DataFrames into a single Excel file with different sheets.
    """
    print(f"  -> Saving reports to: {output_path.resolve()}")
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            for sheet_name, df in reports.items():
                if df.empty:
                    continue
                write_index = "Per_Class" in sheet_name
                df.to_excel(writer, sheet_name=sheet_name, index=write_index)
                if openpyxl:
                    ws = writer.sheets[sheet_name]
                    percent_cols = [
                        c
                        for c in df.columns
                        if c not in ["TP", "FP", "FN", "TN", "Support"]
                    ]
                    for col_name in percent_cols:
                        if col_name in df.columns:
                            col_idx = (
                                df.columns.get_loc(col_name)
                                + (1 if write_index else 0)
                                + 1
                            )
                            for cell in ws[get_column_letter(col_idx)][1:]:
                                cell.number_format = "0.00%"
    except Exception as e:
        print(
            f"  --> ERROR: Could not write to Excel file '{output_path.name}'. Details: {e}"
        )


def main():
    """
    Main function: gestisce le due modalità (directory o singolo file)
    e usa la struttura Models/Model_X/<Dataset>/parser_output_prompt_X.
    """
    parser = argparse.ArgumentParser(
        description="Generates classification metric reports for models or a single file.",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument(
        "models_base_directory",
        type=Path,
        nargs="?",
        default=None,
        help="The base directory containing model subfolders (e.g., 'Models/').\nRequired if --file is not used.",
    )
    parser.add_argument(
        "-f",
        "--file",
        type=Path,
        default=None,
        help="Process a single Excel file instead of a directory.\nOverrides the directory argument.",
    )
    parser.add_argument(
        "-x",
        "--cwe-xml",
        type=Path,
        default=Path("cwec_latest.xml"),
        help="Path to the official CWE XML for validation.\nIf not found, it will be downloaded. (Default: cwec_latest.xml)",
    )
    args = parser.parse_args()

    if not args.file and not args.models_base_directory:
        parser.error(
            "You must provide either a base directory or a single file with -f/--file."
        )

    # Setup iniziale: XML CWE
    if not ensure_cwe_xml_exists(args.cwe_xml):
        sys.exit(1)
    valid_cwe_ids_from_xml = parse_cwe_xml_for_ids(args.cwe_xml)
    if valid_cwe_ids_from_xml is None:
        sys.exit(1)

    # --- SINGLE FILE MODE ---
    if args.file:
        print("\n" + "=" * 80 + "\n--- Running in Single File Mode ---\n" + "=" * 80)
        input_file = args.file
        if not input_file.is_file():
            print(f"ERROR: The specified file does not exist: {input_file}")
            sys.exit(1)

        print(f"\n[+] Processing file: {input_file}")
        filtered_data = load_and_filter_data(input_file, valid_cwe_ids_from_xml)

        if filtered_data is not None and not filtered_data.empty:
            print(
                "  -> Filtering predicted CWEs to only include those present in the ground truth vocabulary..."
            )
            ground_truth_vocabulary = set(
                cwe for cwe_list in filtered_data[COL_GROUND_TRUTH] for cwe in cwe_list
            )
            filtered_data[COL_PREDICTED] = filtered_data[COL_PREDICTED].apply(
                lambda predicted_list: [
                    cwe for cwe in predicted_list if cwe in ground_truth_vocabulary
                ]
            )
            print(
                f"  -> Ground truth vocabulary contains {len(ground_truth_vocabulary)} unique classes."
            )

            per_class, macro_summary, weighted_summary = generate_metric_reports(
                filtered_data
            )
            if not per_class.empty:
                # Path atteso: Models/Model_A/Sven/parser_output_prompt_1/file.xlsx
                prompt_dir = input_file.parent          # .../Sven/parser_output_prompt_1
                dataset_dir = prompt_dir.parent         # .../Sven
                model_dir = dataset_dir.parent          # .../Model_A

                dataset_name = dataset_dir.name
                model_name = model_dir.name

                prompt_number_match = re.search(r'\d+$', prompt_dir.name)
                prompt_number = (
                    prompt_number_match.group(0) if prompt_number_match else "X"
                )

                # Output in: Models/Model_A/Sven/metrics-scenario-2/
                output_dir = dataset_dir / "metrics-scenario-2"
                output_dir.mkdir(parents=True, exist_ok=True)
                print(f"  -> Report will be saved in: {output_dir.resolve()}")

                output_filename = f"metrics_2_prompt_{prompt_number}_{model_name}.xlsx"
                output_xlsx_path = output_dir / output_filename
                reports_to_save = {
                    "Per_Class_Metrics": per_class,
                    "Macro_Avg_Metrics": macro_summary,
                    "Weighted_Avg_Metrics": weighted_summary,
                }
                save_reports_to_excel(output_xlsx_path, reports=reports_to_save)
            else:
                print(
                    "  -> Metrics report was not generated (no valid classes to compare)."
                )
        else:
            print(f"  -> Invalid data or empty file, skipping {input_file.name}")

    # --- DIRECTORY SCAN MODE ---
    else:
        print("\n" + "=" * 80 + "\n--- Running in Directory Scan Mode ---\n" + "=" * 80)
        models_base_dir = args.models_base_directory
        if not models_base_dir.is_dir():
            print(
                f"ERROR: The specified base directory does not exist: {models_base_dir}"
            )
            sys.exit(1)

        print(f"--- Starting Scan in Base Directory: {models_base_dir.resolve()} ---\n")
        for model_dir in models_base_dir.iterdir():
            if not model_dir.is_dir():
                continue

            print("=" * 80 + f"\nProcessing Model: {model_dir.name}\n" + "=" * 80)

            # Per ogni dataset (Sven, PrimeVul, DiverseVul)
            for dataset_name in DATASETS:
                dataset_dir = model_dir / dataset_name
                if not dataset_dir.is_dir():
                    print(
                        f"  INFO: Dataset directory '{dataset_name}' not found for model {model_dir.name}. Skipping."
                    )
                    continue

                output_dir = dataset_dir / "metrics-scenario-2"
                output_dir.mkdir(parents=True, exist_ok=True)
                print(f"\n  Dataset: {dataset_name}")
                print(
                    f"  Reports for this dataset will be saved in: {output_dir.resolve()}\n"
                )

                processed_files_count = 0

                for i in range(1, 4):
                    prompt_subdir = dataset_dir / f"parser_output_prompt_{i}"
                    if not prompt_subdir.is_dir():
                        print(
                            f"  INFO: Subdirectory {prompt_subdir} not found. Skipping."
                        )
                        continue

                    print(
                        f"  --- Scanning prompt directory: {prompt_subdir.name} ---"
                    )
                    for input_xlsx_path in prompt_subdir.glob("*.xlsx"):
                        print(f"\n  [+] Processing file: {input_xlsx_path.name}")
                        filtered_data = load_and_filter_data(
                            input_xlsx_path, valid_cwe_ids_from_xml
                        )
                        if filtered_data is not None and not filtered_data.empty:
                            print(
                                "    -> Filtering predicted CWEs to only include those present in the ground truth vocabulary..."
                            )
                            ground_truth_vocabulary = set(
                                cwe
                                for cwe_list in filtered_data[COL_GROUND_TRUTH]
                                for cwe in cwe_list
                            )
                            filtered_data[COL_PREDICTED] = filtered_data[
                                COL_PREDICTED
                            ].apply(
                                lambda predicted_list: [
                                    cwe
                                    for cwe in predicted_list
                                    if cwe in ground_truth_vocabulary
                                ]
                            )
                            print(
                                f"    -> Ground truth vocabulary contains {len(ground_truth_vocabulary)} unique classes."
                            )

                            per_class, macro_summary, weighted_summary = (
                                generate_metric_reports(filtered_data)
                            )
                            if not per_class.empty:
                                output_filename = (
                                    f"metrics_2_prompt_{i}_{model_dir.name}.xlsx"
                                )
                                output_xlsx_path = output_dir / output_filename

                                reports_to_save = {
                                    "Per_Class_Metrics": per_class,
                                    "Macro_Avg_Metrics": macro_summary,
                                    "Weighted_Avg_Metrics": weighted_summary,
                                }
                                save_reports_to_excel(
                                    output_xlsx_path, reports=reports_to_save
                                )
                                processed_files_count += 1
                            else:
                                print("    -> Metrics report was not generated.")
                        else:
                            print(
                                f"    -> Invalid data or empty file, skipping {input_xlsx_path.name}"
                            )

                print(
                    f"\n  --- Processing for model {model_dir.name}, dataset {dataset_name} complete. "
                    f"Total files processed: {processed_files_count} ---\n"
                )

        print("\n" + "=" * 80 + "\nProcessing complete.\n" + "=" * 80)


if __name__ == "__main__":
    main()
