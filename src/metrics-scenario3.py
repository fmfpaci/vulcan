# --- Core Libraries ---
import pandas as pd
import argparse
import re
import numpy as np
import sys
from pathlib import Path
from typing import List, Optional, Tuple, Dict, Set
import xml.etree.ElementTree as ET
import io
import zipfile
from collections import defaultdict

# --- Optional Libraries for extra features ---
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("--> WARNING: 'openpyxl' not found. Advanced Excel formatting will be disabled. "
          "Install with: pip install openpyxl")
    openpyxl = None
    get_column_letter = None

try:
    import requests
except ImportError:
    print("--> FATAL ERROR: 'requests' not found. It is required for XML download. "
          "Install with: pip install requests")
    requests = None

# --- Configuration Constants ---
COL_FILENAME = 'File Name'
COL_PREDICTED = 'Found CWE'
COL_GROUND_TRUTH = 'Actual CWE'
# New column name for the transformed ground truth
COL_HIERARCHICAL_GT = 'Hierarchical Ground Truth'
CWE_LATEST_URL = "https://cwe.mitre.org/data/xml/cwec_latest.xml.zip"

# Dataset attesi nella struttura Models/Model_X/<Dataset>/parser_output_prompt_X
DATASETS = ["Sven", "PrimeVul", "DiverseVul"]

# ==============================================================================
# --- SECTION 1: INPUT FILE HANDLING AND PARSING (XML) ---
# ==============================================================================

def ensure_cwe_xml_exists(xml_path: Path) -> bool:
    """
    Checks if the CWE XML file exists. If not, it attempts to download and extract it.
    """
    if xml_path.is_file():
        print(f"  -> Found existing CWE XML file: {xml_path}")
        return True
    if not requests:
        return False
    print(f"--- CWE XML file not found at '{xml_path}'. Attempting to download... ---")
    try:
        response = requests.get(CWE_LATEST_URL, timeout=60)
        response.raise_for_status()
        with zipfile.ZipFile(io.BytesIO(response.content)) as z:
            xml_name = next((name for name in z.namelist() if name.endswith('.xml')), None)
            if not xml_name:
                print("  --> FATAL ERROR: No XML file found in the downloaded archive.")
                return False
            print(f"  -> Extracting '{xml_name}' to '{xml_path}'...")
            xml_path.parent.mkdir(parents=True, exist_ok=True)
            with z.open(xml_name) as source, open(xml_path, 'wb') as target:
                target.write(source.read())
        print("  -> Successfully downloaded and extracted the CWE XML file.")
        return True
    except Exception as e:
        print(f"  --> FATAL ERROR during XML download or extraction: {e}")
        return False

def parse_cwe_xml_for_ids(xml_file_path: Path) -> Optional[Set[str]]:
    """
    Parses the CWE XML to extract a set of all valid CWE IDs for validation.
    """
    print(f"  -> Parsing XML to get all valid CWE IDs from: {xml_file_path.name}")
    try:
        tree = ET.parse(xml_file_path)
        root = tree.getroot()
        ns = {'cwe': 'http://cwe.mitre.org/cwe-7'}  # XML namespace is crucial
        cwe_ids = set()
        for element_type in ['Weaknesses', 'Categories', 'Views']:
            container = root.find(f'cwe:{element_type}', ns)
            if container is not None:
                for item in container:
                    if item.get('ID'):
                        cwe_ids.add(item.get('ID'))
        if not cwe_ids:
            print("  --> WARNING: Could not extract any CWE IDs from the XML.")
            return None
        print(f"  -> Found {len(cwe_ids)} unique CWE entries in the XML.")
        return cwe_ids
    except (ET.ParseError, FileNotFoundError) as e:
        print(f"  --> FATAL ERROR: Could not read or parse XML file '{xml_file_path}'.\n      Details: {e}")
        return None

def generate_parent_child_maps_from_xml(xml_path: Path) -> Tuple[Optional[Dict[str, List[str]]], Optional[Dict[str, List[str]]]]:
    """
    Parses the CWE XML to build maps of child-to-parent and parent-to-child relationships.
    """
    print("  -> Generating in-memory CWE parent and child maps from XML...")
    try:
        namespace = {'cwe': 'http://cwe.mitre.org/cwe-7'}
        root = ET.parse(xml_path).getroot()
        child_to_parents_map = defaultdict(list)
        parent_to_children_map = defaultdict(list)
        element_map = {'Weaknesses': 'Weakness', 'Categories': 'Category'}

        for container_tag, element_tag in element_map.items():
            container_xml = root.find(f'cwe:{container_tag}', namespace)
            if container_xml is None:
                continue

            for element in container_xml.findall(f'cwe:{element_tag}', namespace):
                child_id = element.get('ID')
                relations = element.find('cwe:Related_Weaknesses', namespace)
                if relations is not None:
                    for rel in relations.findall('cwe:Related_Weakness', namespace):
                        if rel.get('Nature') == 'ChildOf':
                            parent_id = rel.get('CWE_ID')
                            child_to_parents_map[child_id].append(parent_id)
                            parent_to_children_map[parent_id].append(child_id)

        print(f"  -> Parent map created for {len(child_to_parents_map)} CWEs.")
        print(f"  -> Child map created for {len(parent_to_children_map)} CWEs.")
        return dict(child_to_parents_map), dict(parent_to_children_map)
    except (FileNotFoundError, ET.ParseError) as e:
        print(f"  --> ERROR while generating the parent/child maps: {e}", file=sys.stderr)
        return {}, {}

# ==============================================================================
# --- SECTION 2: DATA PROCESSING AND CORE LOGIC ---
# ==============================================================================

def normalize_cwe_tag(tag: str) -> str:
    """
    Normalizes a CWE tag by removing leading zeros from its ID part.
    Example: 'CWE-020' becomes 'CWE-20'.
    """
    tag_upper = tag.upper()
    if not tag_upper.startswith('CWE-'):
        return tag_upper  # 'NOT VULNERABLE' or others

    match = re.search(r'\d+', tag_upper)
    if match:
        cwe_number = int(match.group(0))
        return f"CWE-{cwe_number}"

    return tag_upper

def process_cwe_tags(series: pd.Series) -> List[str]:
    """
    Extracts and normalizes CWE tags from text.
    """
    if series.dropna().empty:
        return []
    full_text = ' '.join(series.dropna().astype(str))
    tags = re.findall(r'CWE-\d+|NOT VULNERABLE', full_text, re.IGNORECASE)
    normalized_tags = {normalize_cwe_tag(tag) for tag in tags}
    return sorted(list(normalized_tags))

def find_parents_and_children_for_one(cwe_id: str, parent_map: Dict[str, List[str]], child_map: Dict[str, List[str]]) -> List[str]:
    """
    Finds the immediate parents and children for a single CWE ID.
    """
    related_ids: Set[str] = set()
    related_ids.add(cwe_id)

    parents = parent_map.get(cwe_id, [])
    related_ids.update(parents)

    children = child_map.get(cwe_id, [])
    related_ids.update(children)

    return sorted(list(related_ids), key=int)

def apply_new_hierarchy_logic(cwe_list: List[str], parent_map: Dict[str, List[str]], child_map: Dict[str, List[str]]) -> List[str]:
    """
    Applies the parent/child-finding logic to a list of CWE tags.
    """
    final_tags: Set[str] = set()
    for tag_str in cwe_list:
        if tag_str.upper() == 'NOT VULNERABLE':
            final_tags.add('NOT VULNERABLE')
            continue

        cwe_id_match = re.search(r'\d+', tag_str)
        if not cwe_id_match:
            continue

        cwe_id = cwe_id_match.group(0)
        related_cwes = find_parents_and_children_for_one(cwe_id, parent_map, child_map)

        for related_id in related_cwes:
            final_tags.add(f"CWE-{related_id}")

    return sorted(list(final_tags))

def load_and_process_data(
    file_path: Path,
    parent_map: Dict[str, List[str]],
    child_map: Dict[str, List[str]],
    valid_cwe_ids_from_xml: Set[str]
) -> Optional[pd.DataFrame]:
    """
    Loads, filters, and transforms the data from a single Excel file.
    """
    print(f"  -> Reading and processing file: {file_path.name}")
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        required_cols = {COL_FILENAME, COL_PREDICTED, COL_GROUND_TRUTH}
        if not required_cols.issubset(df.columns):
            print(f"  --> ERROR: Missing columns in {file_path.name}! Must contain: {', '.join(required_cols)}")
            return None
    except Exception as e:
        print(f"  --> ERROR: Could not read Excel file {file_path.name}. Details: {e}")
        return None

    aggregated_df = df.groupby(COL_FILENAME).agg({
        COL_GROUND_TRUTH: process_cwe_tags,
        COL_PREDICTED: process_cwe_tags
    }).reset_index()

    print("  -> Filtering 'Found CWE' using official XML data (if applicable)...")
    discard_report = []
    updated_predicted_cwe = []

    for _, row in aggregated_df.iterrows():
        filename, actual_cwes, original_predicted_cwes = row[COL_FILENAME], row[COL_GROUND_TRUTH], row[COL_PREDICTED]
        is_gt_not_vulnerable = (len(actual_cwes) == 1 and actual_cwes[0] == 'NOT VULNERABLE')

        if is_gt_not_vulnerable:
            updated_predicted_cwe.append(original_predicted_cwes)
        else:
            kept_cwes, discarded_cwes = [], []
            for tag in original_predicted_cwes:
                if tag.upper() == 'NOT VULNERABLE':
                    kept_cwes.append(tag)
                    continue
                match = re.search(r'\d+', tag)
                if match and str(int(match.group(0))) in valid_cwe_ids_from_xml:
                    kept_cwes.append(tag)
                else:
                    discarded_cwes.append(tag)
            updated_predicted_cwe.append(kept_cwes)
            if discarded_cwes:
                discard_report.append({"filename": filename, "discarded": sorted(list(set(discarded_cwes)))})

    aggregated_df[COL_PREDICTED] = updated_predicted_cwe

    if discard_report:
        print("\n" + "-"*50)
        print(f"  Summary of invalid CWEs discarded from 'Found CWE' in: {file_path.name}")
        print("-"*50)
        for item in discard_report:
            print(f"  - From row '{item['filename']}': Discarded: {', '.join(item['discarded'])}")
        print("-"*50 + "\n")

    print("  -> Mapping Ground Truth to its parents and children (tree logic)...")
    aggregated_df[COL_HIERARCHICAL_GT] = aggregated_df[COL_GROUND_TRUTH].apply(
        lambda cwes: apply_new_hierarchy_logic(cwes, parent_map, child_map)
    )

    print(f"  -> Metrics will be calculated by comparing '{COL_PREDICTED}' against '{COL_HIERARCHICAL_GT}'.")
    return aggregated_df

# ==============================================================================
# --- SECTION 3: METRICS CALCULATION AND REPORTING ---
# ==============================================================================

def generate_metric_reports(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Calculates classification metrics based on the intersection of two sets of CWEs.
    """
    print(f"  -> Calculating metrics based on set intersection ('{COL_HIERARCHICAL_GT}' vs. '{COL_PREDICTED}')...")

    tp_counts = defaultdict(int)
    fp_counts = defaultdict(int)
    fn_counts = defaultdict(int)
    all_cwes_in_gt = set()
    support_counts = defaultdict(int)
    all_cwes: Set[str] = set()

    for _, row in df.iterrows():
        gt_set = set(cwe for cwe in row[COL_GROUND_TRUTH] if cwe)
        all_cwes_in_gt.update(gt_set)

    for _, row in df.iterrows():
        preds = set(row[COL_PREDICTED])
        gts = set(row[COL_HIERARCHICAL_GT])
        grd = set(row[COL_GROUND_TRUTH])
        all_cwes.update(grd)

        is_not_vulnerable_gt = 'NOT VULNERABLE' in gts
        is_not_vulnerable_pred = 'NOT VULNERABLE' in preds

        if is_not_vulnerable_gt:
            support_counts['NOT VULNERABLE'] += 1
            if is_not_vulnerable_pred:
                tp_counts['NOT VULNERABLE'] += 1
            else:
                fn_counts['NOT VULNERABLE'] += 1
                intersect = preds.intersection(all_cwes_in_gt)
                if len(intersect) > 0:
                    for cwe in intersect:
                        fp_counts[cwe] += 1
        else:
            if is_not_vulnerable_pred:
                fp_counts['NOT VULNERABLE'] += 1
                for cwe in grd:
                    fn_counts[cwe] += 1
            else:
                intersection = preds.intersection(gts)
                if len(intersection) > 0:
                    for cwe in grd:
                        tp_counts[cwe] += 1
                else:
                    for cwe in grd:
                        fn_counts[cwe] += 1

                common = preds.intersection(all_cwes_in_gt)
                for cwe in common - grd:
                    fp_counts[cwe] += 1

            for cwe in grd:
                support_counts[cwe] += 1

    per_class_df = pd.DataFrame(index=sorted(list(all_cwes_in_gt)))
    per_class_df.index.name = 'Class'
    per_class_df['TP'] = per_class_df.index.map(tp_counts).fillna(0).astype(np.int64)
    per_class_df['FP'] = per_class_df.index.map(fp_counts).fillna(0).astype(np.int64)
    per_class_df['FN'] = per_class_df.index.map(fn_counts).fillna(0).astype(np.int64)
    per_class_df['Support'] = per_class_df.index.map(support_counts).fillna(0).astype(np.int64)

    total_samples = len(df)
    per_class_df['TN'] = total_samples - (per_class_df['TP'] + per_class_df['FP'] + per_class_df['FN'])
    per_class_df['TN'] = per_class_df['TN'].apply(lambda x: max(0, x))

    with np.errstate(divide='ignore', invalid='ignore'):
        per_class_df['Precision'] = (per_class_df['TP'] / (per_class_df['TP'] + per_class_df['FP'])).fillna(0)
        per_class_df['Recall'] = (per_class_df['TP'] / per_class_df['Support']).fillna(0)
        per_class_df['F1-Score'] = (2 * per_class_df['Precision'] * per_class_df['Recall'] /
                                    (per_class_df['Precision'] + per_class_df['Recall'])).fillna(0)
        per_class_df['FPR'] = (per_class_df['FP'] / (per_class_df['FP'] + per_class_df['TN'])).fillna(0)
        per_class_df['FNR'] = (per_class_df['FN'] / per_class_df['Support']).fillna(0)

    avg_cols = ['F1-Score', 'FPR', 'FNR']
    final_cols = ['TP', 'FP', 'FN', 'TN', 'Support'] + avg_cols
    per_class_report = per_class_df.reindex(columns=final_cols).sort_index().fillna(0)

    print("  -> Excluding classes with Support=0 from the calculation of general averages.")
    metrics_for_avg = per_class_report[per_class_report['Support'] > 0]
    macro_avg = metrics_for_avg[avg_cols].mean()
    total_support = per_class_report['Support'].sum()
    if total_support > 0:
        weighted_avg = (per_class_report[avg_cols].multiply(per_class_report['Support'], axis=0)).sum() / total_support
    else:
        weighted_avg = pd.Series(0, index=avg_cols)

    return per_class_report, pd.DataFrame(macro_avg).T.add_prefix('Macro '), pd.DataFrame(weighted_avg).T.add_prefix('Weighted ')

def save_reports_to_excel(output_path: Path, reports: Dict[str, pd.DataFrame]):
    """Saves multiple DataFrames into a single Excel file with different sheets."""
    print(f"  -> Saving reports to: {output_path.name}")
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in reports.items():
                if df.empty:
                    print(f"    - Skipping empty sheet: '{sheet_name}'")
                    continue
                write_index = "Per_Class" in sheet_name or "Zero_Support" in sheet_name
                df.to_excel(writer, sheet_name=sheet_name, index=write_index)
                if openpyxl:
                    ws = writer.sheets[sheet_name]
                    percent_cols = [c for c in df.columns if c not in ['TP', 'FP', 'FN', 'TN', 'Support']]
                    for col_name in percent_cols:
                        if col_name in df.columns:
                            col_idx = df.columns.get_loc(col_name) + (1 if write_index else 0) + 1
                            for cell in ws[get_column_letter(col_idx)][1:]:
                                cell.number_format = '0.00%'
    except Exception as e:
        print(f"  --> ERROR: Could not write to Excel file '{output_path.name}'. Details: {e}")

# ==============================================================================
# --- SECTION 4: MAIN FUNCTION AND ARGUMENT PARSING ---
# ==============================================================================

def main():
    """Main function to orchestrate the entire process."""
    parser = argparse.ArgumentParser(
        description="Generates CWE classification metrics with hierarchical GT, "
                    "expecting structure Models/Model_X/<Dataset>/parser_output_prompt_X.",
        formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument(
        "models_base_directory",
        nargs='?',
        type=Path,
        help="Base directory containing model subfolders (e.g., 'Models/'). Not required if --file is used."
    )
    parser.add_argument(
        "-x", "--cwe-xml",
        type=Path,
        default=Path("cwec_latest.xml"),
        help="Path to the CWE XML file. If not found, it will be downloaded. (Default: cwec_latest.xml)"
    )
    parser.add_argument(
        "-f", "--file",
        type=Path,
        help="Process a single Excel file directly, skipping the directory scan. "
             "Expected path: Models/Model_X/<Dataset>/parser_output_prompt_X/file.xlsx"
    )
    parser.add_argument(
        '--debug',
        action='store_true',
        help="Saves an intermediate debug file with aggregated and equivalent data."
    )
    args = parser.parse_args()

    if not args.file and not args.models_base_directory:
        parser.error("Either a 'models_base_directory' or a single file with --file is required.")

    execution_issues = []

    # --- Initial Setup (run once) ---
    print("--- Initializing: Preparing CWE data ---")
    if not ensure_cwe_xml_exists(args.cwe_xml):
        sys.exit(1)
    valid_cwe_ids_from_xml = parse_cwe_xml_for_ids(args.cwe_xml)
    if not valid_cwe_ids_from_xml:
        sys.exit(1)
    parent_map, child_map = generate_parent_child_maps_from_xml(args.cwe_xml)
    if not parent_map or not child_map:
        print("FATAL ERROR: Could not generate the parent/child maps from XML. Exiting.")
        sys.exit(1)

    # --- SINGLE FILE MODE ---
    if args.file:
        print(f"\n--- Single File Mode Activated for: {args.file.resolve()} ---")
        input_xlsx_path = args.file
        if not input_xlsx_path.is_file() or input_xlsx_path.suffix != '.xlsx':
            print(f"ERROR: The path provided with --file is not a valid .xlsx file: {input_xlsx_path}")
            sys.exit(1)

        processed_data = load_and_process_data(input_xlsx_path, parent_map, child_map, valid_cwe_ids_from_xml)

        if processed_data is None:
            execution_issues.append(f"Failed to process file due to read/format error: {input_xlsx_path.name}")
        elif not processed_data.empty:
            per_class, macro, weighted = generate_metric_reports(processed_data)
            if not per_class.empty:
                # Path atteso: Models/Model_X/<Dataset>/parser_output_prompt_X/file.xlsx
                prompt_dir = input_xlsx_path.parent          # .../<Dataset>/parser_output_prompt_X
                dataset_dir = prompt_dir.parent              # .../<Dataset>
                model_dir = dataset_dir.parent               # .../Model_X

                model_name = model_dir.name
                dataset_name = dataset_dir.name

                prompt_number_match = re.search(r'\d+$', prompt_dir.name)
                prompt_number = prompt_number_match.group(0) if prompt_number_match else "X"

                # Output in: Models/Model_X/<Dataset>/metrics-scenario-3
                output_dir = dataset_dir / "metrics-scenario-3"
                output_dir.mkdir(parents=True, exist_ok=True)
                print(f"Reports for this model and dataset will be saved in: {output_dir.resolve()}\n")

                if args.debug:
                    debug_path = output_dir / f"debug_prompt_{prompt_number}_{model_name}.xlsx"
                    debug_cols = [COL_FILENAME, COL_GROUND_TRUTH, COL_HIERARCHICAL_GT, COL_PREDICTED]
                    debug_df = processed_data[debug_cols].copy()
                    for col in debug_df.columns:
                        if isinstance(debug_df[col].iloc[0], list):
                            debug_df[col] = debug_df[col].apply(lambda x: ';'.join(map(str, x)))
                    debug_df.to_excel(debug_path, index=False)

                output_filename = f"metrics_3_prompt_{prompt_number}_{model_name}.xlsx"
                output_xlsx_path = output_dir / output_filename
                reports_to_save = {
                    'Per_Class_Metrics': per_class[per_class['Support'] > 0],
                    'Macro_Avg_Metrics': macro,
                    'Weighted_Avg_Metrics': weighted
                }
                save_reports_to_excel(output_xlsx_path, reports=reports_to_save)
        else:
            execution_issues.append(f"No data to process after aggregation in file: {input_xlsx_path.name}")

        print("\n" + "="*80 + "\n= EXECUTION SUMMARY & ISSUES (Single File Mode) =\n" + "="*80)
        if not execution_issues:
            print("\n✅ OK: Execution completed successfully.")
        else:
            print(f"\n❌ ATTENTION: Execution completed, but {len(execution_issues)} issue(s) were found:")
            print("-" * 60)
            for i, issue in enumerate(execution_issues, 1):
                print(f"  {i}. {issue}")
            print("-" * 60)
        print("\n" + "="*80)
        sys.exit(0)

    # --- DIRECTORY MODE ---
    models_base_dir = args.models_base_directory
    if not models_base_dir.is_dir():
        print(f"ERROR: The specified directory does not exist: {models_base_dir}")
        sys.exit(1)

    print(f"\n--- Starting Scan in Base Directory: {models_base_dir.resolve()} ---")

    for model_dir in sorted(models_base_dir.iterdir()):
        if not model_dir.is_dir():
            continue

        print("\n" + "="*80 + f"\nProcessing Model: {model_dir.name}\n" + "="*80)

        # Per ogni dataset (Sven, PrimeVul, DiverseVul)
        for dataset_name in DATASETS:
            dataset_dir = model_dir / dataset_name
            if not dataset_dir.is_dir():
                issue_msg = f"Dataset directory '{dataset_name}' not found for model '{model_dir.name}'"
                print(f"  INFO: {issue_msg}. Skipping.")
                execution_issues.append(issue_msg)
                continue

            output_dir = dataset_dir / "metrics-scenario-3"
            output_dir.mkdir(parents=True, exist_ok=True)
            print(f"\n  Dataset: {dataset_name}")
            print(f"  Reports for this dataset will be saved in: {output_dir.resolve()}\n")

            for i in range(1, 4):
                prompt_subdir = dataset_dir / f"parser_output_prompt_{i}"
                if not prompt_subdir.is_dir():
                    issue_msg = f"Prompt directory not found for model '{model_dir.name}', dataset '{dataset_name}': {prompt_subdir.name}"
                    print(f"    INFO: {issue_msg}. Skipping.")
                    execution_issues.append(issue_msg)
                    continue

                print(f"    --- Scanning prompt directory: {prompt_subdir.name} ---")
                input_xlsx_files = list(prompt_subdir.glob("*.xlsx"))
                if not input_xlsx_files:
                    issue_msg = f"No Excel file (.xlsx) found in: '{prompt_subdir}'"
                    print(f"      -> WARNING: {issue_msg}. Skipping.")
                    execution_issues.append(issue_msg)
                    continue

                if len(input_xlsx_files) > 1:
                    issue_msg = f"Multiple .xlsx files found in '{prompt_subdir}'. Processing only '{input_xlsx_files[0].name}'"
                    print(f"      -> WARNING: {issue_msg}")
                    execution_issues.append(issue_msg)

                input_xlsx_path = input_xlsx_files[0]
                print(f"\n      [+] Processing: {input_xlsx_path.name}")

                processed_data = load_and_process_data(input_xlsx_path, parent_map, child_map, valid_cwe_ids_from_xml)

                if processed_data is None:
                    issue_msg = f"Failed to process file due to read/format error: {input_xlsx_path.name}"
                    print(f"      -> ERROR: {issue_msg}. Skipping.")
                    execution_issues.append(issue_msg)
                    continue

                if not processed_data.empty:
                    if args.debug:
                        debug_path = output_dir / f"debug_prompt_{i}_{model_dir.name}.xlsx"
                        debug_cols = [COL_FILENAME, COL_GROUND_TRUTH, COL_HIERARCHICAL_GT, COL_PREDICTED]
                        debug_df = processed_data[debug_cols].copy()
                        for col in debug_df.columns:
                            if isinstance(debug_df[col].iloc[0], list):
                                debug_df[col] = debug_df[col].apply(lambda x: ';'.join(map(str, x)))
                        debug_df.to_excel(debug_path, index=False)

                    per_class, macro, weighted = generate_metric_reports(processed_data)
                    if not per_class.empty:
                        output_filename = f"metrics_3_prompt_{i}_{model_dir.name}.xlsx"
                        output_xlsx_path = output_dir / output_filename
                        reports_to_save = {
                            'Per_Class_Metrics': per_class[per_class['Support'] > 0],
                            'Macro_Avg_Metrics': macro,
                            'Weighted_Avg_Metrics': weighted
                        }
                        save_reports_to_excel(output_xlsx_path, reports=reports_to_save)
                else:
                    issue_msg = f"No data to process after aggregation in file: {input_xlsx_path.name}"
                    print(f"      -> NOTICE: {issue_msg}")
                    execution_issues.append(issue_msg)

    print("\n" + "="*80)
    print("= EXECUTION SUMMARY & ISSUES =")
    print("="*80)

    if not execution_issues:
        print("\n✅ OK: Execution completed successfully.")
        print("   All expected files and folders were processed without errors or warnings.")
    else:
        print(f"\n❌ ATTENTION: Execution completed, but {len(execution_issues)} issue(s) were found:")
        print("-" * 60)
        for i, issue in enumerate(execution_issues, 1):
            print(f"  {i}. {issue}")
        print("-" * 60)
        print("   Please check the logs above for detailed context.")

    print("\n" + "="*80)

if __name__ == '__main__':
    main()
